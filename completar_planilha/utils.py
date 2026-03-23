from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX


def identificar_extensao_arquivo():
    ...


def carregar_planilha(caminho_arquivo):
    return load_workbook(caminho_arquivo, keep_vba=True)


def indexed_to_rgb(index):
    try:
        return COLOR_INDEX[index]
    except IndexError:
        return None


def normalize_rgb(rgb):
    """
    Converte 'AARRGGBB' → 'RRGGBB'
    """
    if len(rgb) == 8:
        return rgb[2:]
    return rgb


def obter_aba_exemplo(abas_planilhas: list) -> str | None:
    for aba in abas_planilhas:
        if "exemplo" in aba.lower():
            return aba
    return None


def definir_tipo_lancamento(cor_tupla: tuple | dict | None) -> str:
    """
    Identifica se é 'saida' (cor avermelhada) ou 'entrada'.
    Retorna 'entrada' por padrão caso a cor não seja identificada ou não seja avermelhada.
    """
    if not isinstance(cor_tupla, tuple) or len(cor_tupla) != 3:
        return "entrada"
        
    r, g, b = cor_tupla
    
    # Se o vermelho for predominante, caracteriza como saída
    if r > g and r > b:
        return "saida"
    else:
        return "entrada"


def obter_dados_aba(aba):
    return list(aba.iter_rows(values_only=True))


def criar_dict_referencia(dados_aba, aba_referencia):
    dict_referencia = {}

    for i, linha in enumerate(dados_aba[1:], start=2):
        historico = linha[6].strip()
        if historico in dict_referencia:
            continue
        
        celula = aba_referencia[f"B{i}"]
        
        cor_tupla = obter_valores_rgb_celula(celula)
        tipo_lancamento = definir_tipo_lancamento(cor_tupla)
        tipo_oposto = "entrada" if tipo_lancamento == "saida" else "saida"
        
        deb = linha[3]
        cred = linha[4]

        dict_referencia[historico] = {
            tipo_lancamento: {"deb": deb, "cred": cred},
            tipo_oposto: {"deb": cred, "cred": deb},
        }
    return dict_referencia


def obter_valores_rgb_celula(celula):
    color = celula.font.color

    if color is None:
        return None

    if color.type == "rgb" and color.rgb:
        rgb_normalizado = normalize_rgb(color.rgb)
        rgb_tupla = hex_to_rgb(rgb_normalizado)
        return rgb_tupla

    if color.type == "indexed" and color.indexed is not None:
        try:
            rgb = COLOR_INDEX[color.indexed]
            rgb_normalizado = normalize_rgb(rgb)
            rgb_tupla = hex_to_rgb(rgb_normalizado)
            return rgb_tupla
        except IndexError:
            return None

    # --- Theme (limitado) ---
    if color.type == "theme":
        # Aqui não tem RGB direto sem acessar o tema do workbook
        return {
            "type": "theme",
            "theme": color.theme,
            "tint": color.tint,
        }

    return None


def hex_to_rgb(hex_str):
    """
    Converte uma string hexadecimal 'RRGGBB' em uma tupla (R, G, B).
    """
    if not hex_str or not isinstance(hex_str, str):
        return None
    
    # Remove o '#' se existir
    hex_str = hex_str.lstrip('#')
    
    # Garante que temos 6 caracteres para converter
    if len(hex_str) == 6:
        r = int(hex_str[0:2], 16)
        g = int(hex_str[2:4], 16)
        b = int(hex_str[4:6], 16)
        return (r, g, b)
        
    return None


def obter_aba_atual(abas_planilhas:list) -> str | None:
    for aba in abas_planilhas:
        minuscula_aba = aba.lower()
        if "extrato" in minuscula_aba and "2026" in minuscula_aba:
            return aba
    return None


def preencher_dados(dados_aba, dict_referencia, aba_atual):

    for i, linha in enumerate(dados_aba[1:], start=2):
        historico = linha[5].strip()
        
        if dict_referencia.get(historico):
            celula = aba_atual[f"B{i}"]
            
            cor_tupla = obter_valores_rgb_celula(celula)
            tipo_lancamento = definir_tipo_lancamento(cor_tupla)
            
            aba_atual.cell(row=i, column=3).value = dict_referencia[historico][tipo_lancamento]["deb"]
            aba_atual.cell(row=i, column=4).value = dict_referencia[historico][tipo_lancamento]["cred"]
